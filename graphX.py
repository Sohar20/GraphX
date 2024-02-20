# -*- coding: utf-8 -*-
# GraphX is an app for building mathematical graphs and
# graphs with excel data

import tkinter as tk
from sys import exit
from tkinter.messagebox import showerror, showwarning, showinfo
from math import sin, cos, radians
from math import tan as tg
from math import factorial as fact
import matplotlib.pyplot as plt
from tkinter import ttk
from tkinter.colorchooser import askcolor
from openpyxl import workbook, load_workbook
c = [1, 2, 5, 8, 10, 20, 25, 50, 100, 125, 250, 500, 1000]
color_math = color_1 = color_2 = color_3 = '#000000'

menu_vidgets_del = ['error_t', 'back_btn', 'eng_btn', 'rus_btn', 'f_text', 'entry_f', 'rang',
    'st_x', 'rang1', 'en_x', 'info_button', 'curvature_s', 'curvature_t', 'build_b', 'entry_path1',
    'entry_path2', 'entry_path3', 'entry_sheet1', 'entry_sheet2', 'entry_sheet3', 'data_x1',
    'data_x2', 'data_x3', 'data_y1', 'data_y2', 'data_y3',
    'title_gl', 'title_lOX', 'title_lOY', 'data_x1l', 'data_x2l', 'data_x3l',
    'data_y1l', 'data_y2l', 'data_y3l', 'title_line1', 'title_line2', 'title_line3', 'color_math_button', 'title_graph',
    'title_OX', 'title_OY', 'title_path1', 'title_path2', 'title_path3', 'label_g1', 'label_g2', 'label_g3',
    'sheet_l1', 'sheet_l2', 'sheet_l3', 'title_line_l1', 'title_line_l2', 'title_line_l3', 'button_color1', 'button_color2',
    'button_color3', 'ready_button', 'file_premission1', 'file_premission2', 'file_premission3', 
    'wrong_path1', 'wrong_path2', 'wrong_path3', 'wrong_sheet1', 'wrong_sheet2', 'wrong_sheet3', 'wrong_cell1', 'wrong_cell2', 'wrong_cell3']

data_x1 =data_x2=data_x3=data_y1=data_y2=data_y3=title_line1=title_line2 = title_line3= entry_sheet1= \
    entry_sheet2=entry_sheet3=entry_path1=entry_path2=entry_path3=\
        file_premission1=file_premission2=file_premission3=wrong_path1=wrong_path2=\
            wrong_path3=wrong_sheet1=wrong_sheet2=wrong_sheet3=wrong_cell1=wrong_cell2=wrong_cell3=0
def menu_place_forget():
    # deletes (hides) vidgets from main menu
    title.place_forget()
    start_btn.place_forget()
    sett_btn.place_forget()
    exit_btn.place_forget()
    excelg_btn.place_forget()

def get_col(a):
    # gets column number from excel format ('IH' for example)
    res = 0
    for i in range(len(a)):
        res+=(ord(a[-(i+1)])-64)*(26**i)
    return res

def ctg(x):
    return cos(x)/sin(x)
class Graph():
    def __init__(self, func, ranx, curv):
        self.func = func
        self.x1, self.x2 = ranx
        self.curv = curv
        self.lx = []
        self.ly = []
    
    def check_data(self):
        # checks, if every data is ok
        res = True
        try:
            self.x1 = float(self.x1)
            self.x2 = float(self.x2)
            x=1
            d = float(eval(self.func))
            self.curv = int(self.curv)
        except:
            res = False
        if res:
            res = self.x1 <= self.x2
        
        if res:
            res = self.curv in c
        return res
    
    def draw(self):
        # draws graph
        x = self.x1
        self.curv = 1/self.curv
        while True:
            self.lx.append(round(x, 5))
            try:
                self.ly.append(round(float(eval(self.func)), 5))

            except:

                self.lx[-1] = None
                self.ly.append(None)
            x += self.curv
            x = round(x, 5)
            if x >= self.x2:
                self.lx.append(self.x2)
                x=self.x2
                self.ly.append(float(eval(self.func)))
                break
        

        fig, ax = plt.subplots()
        global color_math
        d = self.x2-self.x1
        y1=round(float(eval(self.func)), 5)
        #base(plt.ylim(self.x1, self.x2)
        plt.plot(self.lx, self.ly, color=color_math)  
        plt.axhline(y=0, color='#000000')
        plt.axvline(x=0, color='#000000')
        ax.set_title(languages[lang][14]+self.func, fontsize=24)
        plt.grid()
        plt.show()

class Excel_graph():
    def __init__(self, x, y, color, title, path, sheet):
        if x[0][0] == x[1][0]:
            self.diff_x = 1
            self.same_x=0
        else:
            self.same_x=1
            self.diff_x = 0
        
        if y[0][0] == y[1][0]:
            self.diff_y = 1
            self.same_y=0
        else:
            self.diff_y = 0
            self.same_y=1
        
        self.x = x
        self.y = y
        self.data_x = []
        self.data_y = []
        self.color = color
        self.title = title
        self.wb = load_workbook(rf'{path}', data_only=True)
        self.sheet = self.wb[sheet]
    
    def check_data(self):
        try:
            for i in range(self.x[0][self.diff_x], self.x[1][self.diff_x]+1):
                if self.same_x == 0:
                    self.data_x.append(float(''.join(str(self.sheet.cell(row=self.x[0][0], column=i).value).split())))

                else:
                    self.data_x.append(float(''.join(str(self.sheet.cell(row=i, column=self.x[0][1]).value).split())))
            
            
            for i in range(self.y[0][self.diff_y], self.y[1][self.diff_y]+1):
                if self.same_y == 0:
                    self.data_y.append(float(''.join(str(self.sheet.cell(row=self.y[0][0], column=i).value).split())))
                else:
                    self.data_y.append(float(''.join(str(self.sheet.cell(row=i, column=self.y[0][1]).value).split())))
        except:
            return False
        
        if len(self.data_x) != len(self.data_y):
            return False

        return (self.data_x, self.data_y)
    
    def build(self):
        global ax
        ax.plot(self.data_x, self.data_y, linewidth=3, color=self.color, label=self.title)
        ax.legend(bbox_to_anchor=(1,0), loc="lower left")
    

def base(x, base):
    return x**(1/base)

try:
    with open('lan.txt', 'r') as f:
        lang = f.read()
        
except:
    with open('lan.txt', 'w') as f:
        f.write('eng')
        lang = 'eng'

def Build():
    # builds a mathmatical graph 
    funct = entry_f.get()
    start_x = st_x.get()
    finish_x = en_x.get()
    c_f = curvature_s.get()
    g = Graph(funct, (start_x, finish_x), c_f)
    if g.check_data():
        error_t.place_forget()
        g.draw()
    else:
        error_t.place(x=50, y=250)

def info():
    showinfo(title=languages[lang][10],
             message=languages[lang][11])   

def G_c():
    # sets vigtes for building a mathematical graph
    menu_place_forget()
    back_btn.place(x=20, y=20)
    f_text.place(x=20, y=75)
    entry_f.place(x=250, y=75)
    rang.place(x=20, y=125)
    st_x.place(x=280, y=130, width=75)
    rang1.place(x=355, y=125)
    en_x.place(x=385, y=130, width=75)
    curvature_t.place(x=20, y=175)
    curvature_s.place(x=150, y=180)
    info_button.place(x=500, y=75, width=200, heigh=30)
    build_b.place(x=300, y=500, width=200, heigh=50)
    color_math_button.place(x=400, y=175, width=200, heigh=30)

def get_math_color():
    # sets a color for a mathematical graph
    global color_math
    color_math = askcolor(title="Color Chooser")[1]

# fuctions "get_color1", "get_color2" and "get_color3"
# sets color for a graph with excel data
def get_color1():
    global color_1
    color_1 = askcolor(title="Color Chooser")[1]

def get_color2():
    global color_2
    color_2 = askcolor(title="Color Chooser")[1]

def get_color3():
    global color_3
    color_3 = askcolor(title="Color Chooser")[1]

def excel_g():
    # places vidgets for building graph with excel data
    menu_place_forget()
    back_btn.place(x=20, y=20)
    title_gl.place(x=150, y=30)
    title_graph.place(x=400, y=30)
    title_lOX.place(x=25, y=75)
    title_OX.place(x=400, y=75, width=300)
    title_lOY.place(x=25, y=125)
    title_OY.place(x=400, y=125, width=300)
    ready_button.place(x=600, y=700, width=175)
    for i in range(1, 4):
        j = i-1
        exec(f'''
label_g{i}.place(relx=0.5, y=175+200*j, anchor="c")
title_path{i}.place(x=25, y=225+200*j)
entry_path{i}.place(x=175, y=225+200*j, width=500)
sheet_l{i}.place(x=25, y=255+200*j)
entry_sheet{i}.place(x=80, y=255+200*j, width=185)
data_x{i}l.place(x=180, y=255+200*j)
data_x{i}.place(x=325, y=255+200*j, width=100)
data_y{i}l.place(x=450, y=255+200*j)
data_y{i}.place(x=600, y=255+200*j, width=100)
title_line_l{i}.place(x=25, y=325+200*j)
title_line{i}.place(x=200, y=325+200*j)
button_color{i}.place(x=450, y=325+200*j, width=150, height=30)''')

def exit_c():
    exit()

def Engl_c():
    # sets app language as english
    with open('lan.txt', 'w') as f:
        f.write('eng')
        lang = 'eng'
    showinfo(title="Information", message='Reboot required to change settings')
    exit()
        
def Rusl_c():
    #sets app language as russian
    with open('lan.txt', 'w') as f:
        f.write('rus')
        lang = 'rus'
    showinfo(title="Информация", 
             message="Требуется перезагрузка для изменения настроек")
    exit()

def menu_c():
    # places menu vidgets
    for i in menu_vidgets_del:
        eval(f'{i}.place_forget()')
    title.place(x=300, y=100)
    start_btn.place(x=200, y=200, width=400, heigh=50)
    excelg_btn.place(x=200, y=350, width=400, heigh=50)
    sett_btn.place(x=200, y=500, width=400, heigh=50)
    exit_btn.place(x=200, y=650, width=400, heigh=50)
# "check_data1", "chack_data2" and "check_data3" checks entered data for excel
def check_data1():
    path = entry_path1.get()
    sheet=entry_sheet1.get()
    true_path = true_sheet = true_xrange1 = true_yrange = True
    rangex = data_x1.get()
    rangey = data_y1.get()

    try:
        wb=load_workbook(rf'{path}')
    except:
        true_path = False
    else:

        try:
            sheet = wb[sheet]
        except:
            true_sheet = False
        else:
            datax = get_excel_data(rangex)
            datay = get_excel_data(rangey)
    if path:
        if not (path.endswith('xls') or path.endswith('xlsx')):
            file_premission1.place(relx=0.5, y=300, anchor="c")
            return False
        
        elif not true_path:
            file_premission1.place_forget()
            wrong_path1.place(relx=0.5, y=300, anchor="c")
            return False
        
        elif not true_sheet:
            file_premission1.place_forget()
            wrong_path1.place_forget()
            wrong_sheet1.place(relx=0.5, y=300, anchor="c")
            return False
        
        elif not (datax and datay):
             file_premission1.place_forget()
             wrong_path1.place_forget()
             wrong_sheet1.place_forget()
             wrong_cell1.place(relx=0.5, y=300, anchor="c")
             return False
        
        else:
            file_premission1.place_forget()
            wrong_path1.place_forget()
            wrong_sheet1.place_forget()
            wrong_cell1.place_forget()
            return True
    else:
        file_premission1.place_forget()
        wrong_path1.place_forget()
        wrong_sheet1.place_forget()
        wrong_cell1.place_forget()
        return [None]

def check_data2():
    path = entry_path2.get()
    sheet=entry_sheet2.get()
    true_path = true_sheet = true_xrange1 = true_yrange = True
    rangex = data_x2.get()
    rangey = data_y2.get()

    try:
        wb=load_workbook(rf'{path}')
    except:
        true_path = False
    else:

        try:
            sheet = wb[sheet]
        except:
            true_sheet = False
        else:
            datax = get_excel_data(rangex)
            datay = get_excel_data(rangey)
    if path:
        if not (path.endswith('xls') or path.endswith('xlsx')):
            file_premission2.place(relx=0.5, y=500, anchor="c")
            return False
        
        elif not true_path:
            file_premission2.place_forget()
            wrong_path1.place(relx=0.5, y=500, anchor="c")
            return False
        
        elif not true_sheet:
            file_premission2.place_forget()
            wrong_path2.place_forget()
            wrong_sheet2.place(relx=0.5, y=500, anchor="c")
            return False
        
        elif not (datax and datay):
             file_premission2.place_forget()
             wrong_path2.place_forget()
             wrong_sheet2.place_forget()
             wrong_cell2.place(relx=0.5, y=500, anchor="c")
             return False
        
        else:
            file_premission2.place_forget()
            wrong_path2.place_forget()
            wrong_sheet2.place_forget()
            wrong_cell2.place_forget()
            return True
    else:
        file_premission2.place_forget()
        wrong_path2.place_forget()
        wrong_sheet2.place_forget()
        wrong_cell2.place_forget()
        return [None]

def check_data3():
    path = entry_path3.get()
    sheet=entry_sheet3.get()
    true_path = true_sheet = true_xrange1 = true_yrange = True
    rangex = data_x3.get()
    rangey = data_y3.get()

    try:
        wb=load_workbook(rf'{path}')
    except:
        true_path = False
    else:

        try:
            sheet = wb[sheet]
        except:
            true_sheet = False
        else:
            datax = get_excel_data(rangex)
            datay = get_excel_data(rangey)
    if path:
        if not (path.endswith('xls') or path.endswith('xlsx')):
            file_premission3.place(relx=0.5, y=700, anchor="c")
            return False
        
        elif not true_path:
            file_premission3.place_forget()
            wrong_path3.place(relx=0.5, y=700, anchor="c")
            return False
        
        elif not true_sheet:
            file_premission3.place_forget()
            wrong_path3.place_forget()
            wrong_sheet3.place(relx=0.5, y=700, anchor="c")
            return False
        
        elif not (datax and datay):
             file_premission3.place_forget()
             wrong_path3.place_forget()
             wrong_sheet3.place_forget()
             wrong_cell3.place(relx=0.5, y=700, anchor="c")
             return False
        
        else:
            file_premission3.place_forget()
            wrong_path3.place_forget()
            wrong_sheet3.place_forget()
            wrong_cell3.place_forget()
            return True
    else:
        file_premission3.place_forget()
        wrong_path3.place_forget()
        wrong_sheet3.place_forget()
        wrong_cell3.place_forget()
        return [None]

def build_ex():
    #builds graph with excel data
    fig, ax = plt.subplots()
    graph_title = title_graph.get()
    OX_title = title_OX.get()
    OY_title = title_OY.get()

    cells1 = check_data1()
    cells2 = check_data2()
    cells3 = check_data3()
    if cells1 and cells2 and cells3:
        for i in range(1, 4):
            exec(f'''
if cells{i} != [None]:
    x = get_excel_data(data_x{i}.get())
    y = get_excel_data(data_y{i}.get())
    if x and y:
        g{i} = Excel_graph(x, y, color_{i}, title_line{i}.get(), entry_path{i}.get(), entry_sheet{i}.get())
        if not g{i}.check_data():
            wrong_cell{i}.place(relx=0.5, y=300, anchor="c")
                    
        else:
            print('a')
            ax.plot(g{i}.data_x, g{i}.data_y, linewidth=3, color=g{i}.color, label=g{i}.title)
            wrong_cell{i}.place_forget()
            
    else:
        wrong_cell{i}.place(relx=0.5, y=300, anchor="c")''')
        
        
        ax.legend(bbox_to_anchor=(1,0), loc="lower left")
        ax.set_title(graph_title, fontsize=24)
        plt.xlabel(OX_title)
        plt.ylabel(OY_title)
        plt.grid()
        plt.show()
            

def get_excel_data(a):
    # gets cell format data 
    a = a.upper()
    a = list(a.split(':'))
    if len(a) != 2:
        return 0
    x1 = list(a[0])
    x2 = list(a[1])
    numx1 = ''
    numx2 = ''
    if a[0].isdigit() or a[1].isdigit() or a[0].isalpha() or a[1].isalpha():
        return 0
    for i in range(len(x1)):
        numx1 += x1.pop()
        if str(''.join(x1)).isalpha():
            numx1 = int(''.join(list(reversed(numx1))))
            break
    
    else:
        return 0
    
    for i in range(len(x2)):
        numx2 += x2.pop()
        if str(''.join(x2)).isalpha():
            numx2 = int(''.join(list(reversed(numx2))))
            break
    
    else:
        return 0
    
    x1 = ''.join(x1)
    x2 = ''.join(x2)
    res1 = 0
    res2 = 0
    for i in range(len(x1)):
        res1+=(ord(x1[-(i+1)])-64)*(26**i)
    
    for i in range(len(x2)):
        res2+=(ord(x2[-(i+1)])-64)*(26**i)
    
    if len({numx1, x1, numx2, x2}) != 3:
        return 0
    
    if numx1 > numx2 or res1 > res2:
        return 0
    
    return [[numx1, res1], [numx2, res2]]


def sett_c():
    # places settings vidgets
    menu_place_forget()
    back_btn.place(x=20, y=20) 
    eng_btn.place(x=200, y=20)
    rus_btn.place(x=200, y=170)

   
languages = {'eng': ['Build a mathematical graph', 'Settings of language', 'Exit', 'Back',
                     'English', 'Russian', 'Function of the graph y=',
                     'Value range for x from:', 'to', 'Curvature:', 
                     'Reference', '''x - name of the variable
expressions:
+ addition
- subtraction
* multiplication
** exponentiation
/ division
// integer division
features:
abs() - the absolute value of a number
fact() - factorial
base(a, n) root of number a to the nth degree
there are also functions tg(), sin(), cos()''', 'Build',
                     '''Fields filled out incorrectly
Check:
1) The fact that the formula is written correctly, and only functions are used,
listed in the handbook.
2) that the first number in the value range for x is less than the second.
3) That you are using numbers as range arguments
(without spaces)''', 'Function graph for ', 'Build a graph with excel data', 'Title of the graph',
'Title of horizontal axis', 'Title of vertical axis', 'Data for x', 'Data for y', 'Title of the line', 'Color', 'Path to the file',
'Graph №1', 'Graph №2', 'Graph №3', 'sheet', 'Incorrect file permission', 'Wrong path', 'Wrong sheet', 'Incorrect cell format'],
             'rus': ['Построить математический график', 'Настройки языка', 'Выход', 'Назад',
                     'Ангийский', 'Русский', 'Функция графика y=',
                     'Диапазон значений для x от', 'до', 'Кривизна:', 
                     'Справка', '''x - имя переменной
выражения:
+ сложение
- вычитание
* умножение
** возведение в степень
/ деление
// целочисленное деление
функции:
abs() модуль числа
fact() факториал
base(a, n) корень из числа a n-ной степени
также есть функции tg(), sin(), cos(), ctg()s''', 'Построить',
                     '''Поля заполнены неправильно
Проверьте:
1) То, что формула записана правильно, и использованы только функции,
указанные в справочнике.
2) то, что первое число в диапазоне значения для x меньше второго.
3) То, что вы используете числа в качестве аргументов диапазона
(без пробелов)''', 'График функции для ', 'Построить график с excel данными', 'Название графика',
'Название горизонтальной оси', 'Название вертикальной оси', 'Данные для x', 'Данные для y', 'Название линии', 'Цвет', 'Путь к файлу',
'График №1', 'График №2', 'График №3', 'Лист', 'Неправильное разрешение файла', 'Неправильный путь',
'Неправильный лист', 'Неправильный формат ячеек']}

window = tk.Tk()
window.title("graphX") 
window.geometry('800x900') 
f_text = tk.Label(text = languages[lang][6], font=("Arial", 15))
curvature_t = tk.Label(text=languages[lang][9], font=("Arial", 15))
back_btn = tk.Button(text=languages[lang][3], font=("Arial", 20), 
                     command=menu_c)
eng_btn = tk.Button(text=languages[lang][4], font=("Arial", 20), 
                     command=Engl_c)
rus_btn = tk.Button(text=languages[lang][5], font=("Arial", 20), 
                     command=Rusl_c)

error_t = tk.Label(text=languages[lang][13], font=("Arial", 15),
                   foreground='red')
title = tk.Label(text="GraphX", font=("Arial", 50)) 
title.place(x=300, y=50) 
curvature_s = ttk.Combobox(values=c)
color_math_button = tk.Button(text=languages[lang][22], font=("Arial", 15),
                    command=get_math_color)

entry_f = tk.Entry(font = ('Arial', 15))
rang = tk.Label(text = languages[lang][7], font=("Arial", 15))
rang1 = tk.Label(text = languages[lang][8], font=("Arial", 15))
st_x = tk.Entry()
info_button = tk.Button(text=languages[lang][10], font=("Arial", 15),
                        command=info)
en_x = tk.Entry()

start_btn = tk.Button(text=languages[lang][0], 
                      font=("Arial", 15), command=G_c)
start_btn.place(x=200, y=200, width=400, heigh=50)

excelg_btn = tk.Button(text=languages[lang][15], font=("Arial", 15), command=excel_g)
excelg_btn.place(x=200, y=350, width=400, heigh=50)

sett_btn = tk.Button(text=languages[lang][1], font=("Arial", 20),
                     command=sett_c) 
sett_btn.place(x=200, y=500, width=400, heigh=50) 

exit_btn = tk.Button(text=languages[lang][2], font=("Arial", 20), 
                     command=exit_c) 
exit_btn.place(x=200, y=650, width=400, heigh=50)

build_b = tk.Button(text=languages[lang][12], font=("Arial", 25), command=Build)

#vidgets of building a graph with excel data are below
title_gl = tk.Label(text=languages[lang][16], font=('Arial', 20))
title_lOX = tk.Label(text=languages[lang][17], font=('Arial', 17))
title_lOY = tk.Label(text=languages[lang][18], font=('Arial', 17))

title_graph = tk.Entry(font = ('Arial', 20))
title_OX = tk.Entry(font = ('Arial', 17))
title_OY = tk.Entry(font = ('Arial', 17))

entries = ['data_x1', 'data_x2', 'data_x3', 'data_y1', 'data_y2', 'data_y3', 'title_line1',
           'title_line2', 'title_line3', 'entry_sheet1', 'entry_sheet2', 'entry_sheet3',
             'entry_path1', 'entry_path2', 'entry_path3']

for i in entries:
    exec(f"{i} = tk.Entry(font = ('Arial', 15))")

data_x1l =  tk.Label(text=languages[lang][19], font = ('Arial', 15))
data_x2l = tk.Label(text=languages[lang][19], font = ('Arial', 15))
data_x3l = tk.Label(text=languages[lang][19], font = ('Arial', 15))

data_y1l = tk.Label(text=languages[lang][20], font = ('Arial', 15))
data_y2l = tk.Label(text=languages[lang][20], font = ('Arial', 15))
data_y3l = tk.Label(text=languages[lang][20], font = ('Arial', 15))

title_line_l1 = tk.Label(text=languages[lang][21], font = ('Arial', 15))
title_line_l2 = tk.Label(text=languages[lang][21], font = ('Arial', 15))
title_line_l3 = tk.Label(text=languages[lang][21], font = ('Arial', 15))

button_color1 = tk.Button(text=languages[lang][22], font = ('Arial', 15),
                         command=get_color1)
button_color2 = tk.Button(text=languages[lang][22], font = ('Arial', 15),
                         command=get_color2)
button_color3 = tk.Button(text=languages[lang][22], font = ('Arial', 15),
                         command=get_color3)

title_path1 = tk.Label(text=languages[lang][23], font = ('Arial', 15))
title_path2 = tk.Label(text=languages[lang][23], font = ('Arial', 15))
title_path3 = tk.Label(text=languages[lang][23], font = ('Arial', 15))

label_g1 = tk.Label(text=languages[lang][24], font = ('Arial', 17))
label_g2 = tk.Label(text=languages[lang][25], font = ('Arial', 17))
label_g3 = tk.Label(text=languages[lang][26], font = ('Arial', 17))

sheet_l1 = tk.Label(text=languages[lang][27], font = ('Arial', 15))
sheet_l2 = tk.Label(text=languages[lang][27], font = ('Arial', 15))
sheet_l3 = tk.Label(text=languages[lang][27], font = ('Arial', 15))

ready_button = tk.Button(text=languages[lang][12], font=("Arial",25), 
                         command=build_ex)

#error messages in excel graphs
labels_err =['file_premission1', 'file_premission2', 'file_premission3',
             'wrong_path1', 'wrong_path2', 'wrong_path3', 'wrong_sheet1',
             'wrong_sheet2', 'wrong_sheet3', 'wrong_cell1', 'wrong_cell2', 'wrong_cell3']
i = 28
cnt = 0
for el in labels_err:
    exec(f'{el} = tk.Label(text=languages[lang][{i}], font=("Arial",15),\
                           foreground="red")')
    cnt += 1
    if cnt%3==0:
        i+=1

window.mainloop()
